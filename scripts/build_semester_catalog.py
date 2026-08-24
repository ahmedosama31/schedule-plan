"""Build a versioned semester catalog from the university's mixed source formats.

This is an ingestion tool, not part of the web application's runtime. It accepts
the portal TSV export, a faculty PDF table, and the curated schedule workbook,
then writes deterministic JSON that can be reviewed and committed.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


DAYS = {"Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday"}

# The workbook contains course titles but not codes. These mappings were checked
# against Cairo University Faculty of Engineering's 2023 CCE bylaws and Fall 2025
# exam schedule. Keep this mapping explicit so future updates fail visibly when a
# new or renamed course appears.
WORKBOOK_CODE_BY_NAME = {
    "APT": "CMPS211",
    "Advanced Cybersecurity": "CMPS452",
    "Advanced Database Systems": "CMPS401",
    "Algorithms Design and Analysis": "CMPS302",
    "Comp. Graphics and Real-Time Rendering": "CMPS346",
    "Compilers and Languages": "CMPS403",
    "Computer Architecture": "CMPS301",
    "Computer Consultation": "CMPS425",
    "Computer Networks-1": "CMPS405",
    "Data Structures and Algorithms": "CMPS103",
    "Engineering Seminar": "CCES280",
    "Generative AI for Software Testing": "CMPS457",
    "Graduation Project-1": "CCES481",
    "Image Processing and Computer Vision": "CMPS446",
    "Intro to Database Management Systems": "CMPS202",
    "Introduction to Logic Design": "CMPS118",
    "Logic Design": "CMPS101",
    "Machine Intelligence": "CMPS402",
    "Microprocessor Systems": "CMPS201",
    "Mobile robotics": "CMPS456",
    "Natural Language Processing": "CMPS454",
    "Operating Systems": "CMPS303",
    "Pattern Recognition": "CMPS450",
    "Programming Techniques": "CMPS102",
    "Project Management": "GENS249",
    "Reinforcement Learning Course": "CMPS458",
}

CANONICAL_NAME_BY_CODE = {
    "CMPS211": "Advanced Programming Techniques",
    "CMPS302": "Algorithms Design and Analysis",
    "CMPS346": "Computer Graphics and Real-Time Rendering",
    "CMPS450": "Pattern Recognition and Artificial Neural Networks",
}


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def clean(value: Any) -> str:
    return str(value or "").replace("_", "").strip()


def parse_clock(value: str) -> tuple[int, int]:
    match = re.fullmatch(r"\s*(\d{1,2}):(\d{2})\s*", clean(value))
    if not match:
        raise ValueError(f"Invalid time: {value!r}")
    hour, minute = map(int, match.groups())
    if 1 <= hour <= 7:
        hour += 12
    if not 0 <= hour <= 23 or not 0 <= minute <= 59:
        raise ValueError(f"Invalid time: {value!r}")
    return hour, minute


def time_payload(start: str, end: str, rounded_end: bool = False) -> dict[str, Any]:
    start_hour, start_minute = parse_clock(start)
    end_hour, end_minute = parse_clock(end)
    if end_hour * 60 + end_minute <= start_hour * 60 + start_minute and end_hour < 12:
        end_hour += 12
    if rounded_end:
        # The curated workbook displays block boundaries (e.g. 08:00-11:00),
        # while portal/PDF rows use the actual 10:50 end time.
        end_hour, end_minute = divmod(end_hour * 60 + end_minute - 10, 60)
    return {
        "startHour": start_hour + start_minute / 60,
        "endHour": end_hour + end_minute / 60,
        "startString": f"{start_hour:02d}:{start_minute:02d}",
        "endString": f"{end_hour:02d}:{end_minute:02d}",
    }


def normalize_type(value: str) -> str:
    lowered = clean(value).lower()
    if "lab" in lowered:
        return "Laboratory"
    if "tutorial" in lowered or "section" in lowered:
        return "Tutorial"
    return "Lecture"


def portal_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for line_number, line in enumerate(path.read_text(encoding="utf-8-sig").splitlines(), 1):
        parts = [part.strip() for part in line.split("\t")]
        if len(parts) < 8:
            continue
        code = clean(parts[1])
        day = clean(parts[5])
        if not code or day not in DAYS:
            continue
        session = time_payload(parts[6], parts[7])
        if len(parts) > 12 and parts[12].strip():
            session["location"] = parts[12].strip()
        rows.append({
            "code": code,
            "name": parts[2].strip(),
            "group": clean(parts[3]) or "1",
            "type": normalize_type(parts[4]),
            "day": day,
            "source": "portal-tsv",
            "sourceRow": line_number,
            **session,
        })
    return rows


def pdf_rows(path: Path) -> list[dict[str, Any]]:
    try:
        import pdfplumber
    except ImportError as exc:
        raise RuntimeError("PDF import requires pdfplumber") from exc

    rows: list[dict[str, Any]] = []
    with pdfplumber.open(path) as document:
        for page_number, page in enumerate(document.pages, 1):
            for table in page.extract_tables() or []:
                for row_number, row in enumerate(table, 1):
                    if not row or len(row) < 7 or clean(row[0]).lower() == "code":
                        continue
                    code = clean(row[0])
                    day = clean(row[4])
                    if not code or day not in DAYS:
                        continue
                    rows.append({
                        "code": code,
                        "name": clean(row[1]),
                        "group": clean(row[2]) or "1",
                        "type": normalize_type(row[3]),
                        "day": day,
                        "source": "faculty-pdf",
                        "sourceRow": f"{page_number}:{row_number}",
                        **time_payload(row[5], row[6]),
                    })
    return rows


def workbook_rows(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Workbook import requires openpyxl") from exc

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Day-by-Day Schedule" not in workbook.sheetnames:
        raise ValueError("Workbook is missing the 'Day-by-Day Schedule' sheet")
    sheet = workbook["Day-by-Day Schedule"]
    current_day: str | None = None
    rows: list[dict[str, Any]] = []
    unknown_names: set[str] = set()

    for row_number, values in enumerate(sheet.iter_rows(values_only=True), 1):
        first = clean(values[0] if values else "")
        day_match = re.match(r"^(SATURDAY|SUNDAY|MONDAY|TUESDAY|WEDNESDAY|THURSDAY)\b", first, re.I)
        if day_match:
            current_day = day_match.group(1).title()
            continue
        if current_day is None or len(values) < 5 or not first or first == "Time Slot":
            continue
        time_match = re.fullmatch(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})", first)
        if not time_match:
            continue
        name = clean(values[1])
        code = WORKBOOK_CODE_BY_NAME.get(name)
        if not code:
            unknown_names.add(name)
            continue
        section_code = clean(values[2])
        section_type = "Lecture" if section_code == "Main Section" else normalize_type(clean(values[4]))
        group_match = re.search(r"(\d+)$", section_code)
        group = group_match.group(1) if group_match else "1"
        rows.append({
            "code": code,
            "name": CANONICAL_NAME_BY_CODE.get(code, name),
            "group": group,
            "type": section_type,
            "day": current_day,
            "source": "curated-workbook",
            "sourceRow": row_number,
            **time_payload(time_match.group(1), time_match.group(2), rounded_end=True),
        })

    if unknown_names:
        raise ValueError(f"Workbook contains unmapped course titles: {sorted(unknown_names)}")
    return rows


def meeting_key(row: dict[str, Any], include_type: bool = True) -> tuple[Any, ...]:
    key_parts = [row["code"]]
    if include_type:
        key_parts.extend([row["type"], row["group"]])
    key_parts.extend([row["day"], row["startString"], row["endString"]])
    return tuple(key_parts)


def merge_rows(sources: Iterable[list[dict[str, Any]]]) -> tuple[list[dict[str, Any]], Counter[str]]:
    merged: list[dict[str, Any]] = []
    exact_seen: set[tuple[Any, ...]] = set()
    meeting_seen: dict[tuple[Any, ...], int] = {}
    section_seen: dict[tuple[str, str, str], int] = {}
    stats: Counter[str] = Counter()

    for source_priority, rows in enumerate(sources):
        for row in rows:
            exact_key = meeting_key(row)
            loose_key = meeting_key(row, include_type=False)
            section_key = (row["code"], row["type"], row["group"])
            if exact_key in exact_seen:
                stats[f"{row['source']}:duplicate"] += 1
                continue
            if section_key in section_seen and section_seen[section_key] < source_priority:
                stats[f"{row['source']}:section-conflict"] += 1
                continue
            # A higher-priority source may classify a meeting more accurately
            # (notably graduation projects in the workbook). Do not add a second
            # section solely because the lower-priority source's type differs.
            if loose_key in meeting_seen and meeting_seen[loose_key] < source_priority:
                stats[f"{row['source']}:type-conflict"] += 1
                continue
            exact_seen.add(exact_key)
            meeting_seen.setdefault(loose_key, source_priority)
            section_seen.setdefault(section_key, source_priority)
            merged.append(row)
            stats[f"{row['source']}:accepted"] += 1
    return merged, stats


def build_courses(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    courses: dict[str, dict[str, Any]] = {}
    sections: dict[tuple[str, str, str], dict[str, Any]] = {}

    for row in rows:
        course = courses.setdefault(row["code"], {
            "code": row["code"],
            "name": CANONICAL_NAME_BY_CODE.get(row["code"], row["name"]),
            "sections": [],
            "isMTHS": row["code"].startswith("MTHS"),
        })
        section_key = (row["code"], row["type"], row["group"])
        if section_key not in sections:
            type_slug = {"Lecture": "lec", "Tutorial": "tut", "Laboratory": "lab"}[row["type"]]
            section = {
                "id": f"{row['code']}-{type_slug}-{row['group']}",
                "legacyIds": [],
                "courseCode": row["code"],
                "type": row["type"],
                "group": row["group"],
                "sessions": [],
            }
            sections[section_key] = section
            course["sections"].append(section)
        section = sections[section_key]
        legacy_id = (
            f"{row['code']}-{row['type']}-{row['group']}-{row['day']}-"
            f"{row['startString'].lstrip('0')}-{row['endString'].lstrip('0')}"
        )
        if legacy_id not in section["legacyIds"]:
            section["legacyIds"].append(legacy_id)
        session = {
            "day": row["day"],
            "startHour": row["startHour"],
            "endHour": row["endHour"],
            "startString": row["startString"],
            "endString": row["endString"],
        }
        if row.get("location"):
            session["location"] = row["location"]
        if session not in section["sessions"]:
            section["sessions"].append(session)

    for course in courses.values():
        course["sections"].sort(key=lambda section: (section["type"], section["group"], section["id"]))
        for section in course["sections"]:
            section["sessions"].sort(key=lambda item: (item["day"], item["startHour"], item["endHour"]))
    return sorted(courses.values(), key=lambda course: course["code"])


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--portal", required=True, type=Path)
    parser.add_argument("--pdf", required=True, type=Path)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--metadata", required=True, type=Path)
    parser.add_argument("--semester-id", required=True)
    parser.add_argument("--semester-label", required=True)
    parser.add_argument("--imported-at", default=datetime.now(timezone.utc).isoformat())
    args = parser.parse_args()

    parsed_portal = portal_rows(args.portal)
    parsed_pdf = pdf_rows(args.pdf)
    parsed_workbook = workbook_rows(args.workbook)
    merged, merge_stats = merge_rows([parsed_portal, parsed_pdf, parsed_workbook])
    courses = build_courses(merged)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.metadata.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(courses, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")

    metadata = {
        "semesterId": args.semester_id,
        "semesterLabel": args.semester_label,
        "importedAt": args.imported_at,
        "courseCount": len(courses),
        "sectionCount": sum(len(course["sections"]) for course in courses),
        "sessionCount": sum(len(section["sessions"]) for course in courses for section in course["sections"]),
        "sources": [
            {"name": args.portal.name, "sha256": sha256(args.portal), "parsedRows": len(parsed_portal)},
            {"name": args.pdf.name, "sha256": sha256(args.pdf), "parsedRows": len(parsed_pdf)},
            {"name": args.workbook.name, "sha256": sha256(args.workbook), "parsedRows": len(parsed_workbook)},
        ],
        "mergeStats": dict(sorted(merge_stats.items())),
        "courseCodeReferences": [
            "https://eng.cu.edu.eg/wp-content/uploads/credituser/2015/CUFE_STEP_2023BYLAWS_CCE_V1-R-op.pdf",
            "https://eng.cu.edu.eg/wp-content/uploads/credituser/2015/FA25-ADM_AN_009-final-fall-205-v_5.pdf",
        ],
    }
    args.metadata.write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(json.dumps({
        "courses": metadata["courseCount"],
        "sections": metadata["sectionCount"],
        "sessions": metadata["sessionCount"],
        "mergeStats": metadata["mergeStats"],
    }, indent=2))


if __name__ == "__main__":
    main()
